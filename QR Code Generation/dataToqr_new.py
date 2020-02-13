from openpyxl import *
import pyqrcode,uuid,png
import pandas as pd

wb = load_workbook('sample.xlsx',data_only=True)
ws = wb['Sheet1']
wb2 = Workbook()
ws2 = wb2.create_sheet('final')

guests = {}
keys = ["name","totalAdults","totalKids","totalAdultsArrived","totalKidsArrived","email","qrLink"]

for i in range (ws.min_row,ws.max_row+1):
    namecellref=ws.cell(row=i,column=1)
    totalAdultscellref=ws.cell(row=i,column=3)
    totalKidscellref=ws.cell(row=i,column=4)
    emailcellref=ws.cell(row=i,column=2)
    guest = dict.fromkeys(keys,"")
    guest["name"] = namecellref.value
    guest["totalAdults"] = totalAdultscellref.value
    guest["totalKids"] = totalKidscellref.value
    guest["totalAdultsArrived"] = 0
    guest["totalKidsArrived"] = 0
    guest["email"] = emailcellref.value
    uid = str(uuid.uuid3(uuid.NAMESPACE_DNS,(str(namecellref.value) + str(namecellref.row)+str(namecellref.column) + "Transcend2020byPanda")))
    guest["qrLink"] = '=HYPERLINK(CONCATENATE("./qrcodes/","'+ uid +'",".png"),CONCATENATE("'+ namecellref.value +'"," qrcode link"))'
    guests[uid]=guest

for i in range (ws.min_row,ws.max_row+1):
    uidcellref=ws2.cell(row=i,column=1)
    namecellref=ws2.cell(row=i,column=2)
    emailcellref=ws2.cell(row=i,column=3)
    totalAdultscellref=ws2.cell(row=i,column=4)
    totalKidscellref=ws2.cell(row=i,column=5)
    totalAdultsArrivedcellref=ws2.cell(row=i,column=6)
    totalKidsArrivedcellref=ws2.cell(row=i,column=7)
    qrLinkcellref=ws2.cell(row=i,column=8)
    uids=[*guests]
    uidcellref.value=str(uids[i-1])
    namecellref.value=str(guests[uids[i-1]]["name"])
    emailcellref.value=str(guests[uids[i-1]]["email"])
    totalAdultscellref.value=int(guests[uids[i-1]]["totalAdults"])
    totalKidscellref.value=int(guests[uids[i-1]]["totalKids"])
    totalAdultsArrivedcellref.value=int(guests[uids[i-1]]["totalAdultsArrived"])
    totalKidsArrivedcellref.value=int(guests[uids[i-1]]["totalKidsArrived"])
    qrLinkcellref.value=guests[uids[i-1]]["qrLink"]

wb2.save('final.xlsx')
pd.read_excel('final.xlsx',sheet_name='final').to_csv('final.csv',index=False)
wb3 = load_workbook(filename='final.xlsx',data_only=True)
ws3 = wb3['final']

for cell in ws3['A']:
    qr = pyqrcode.create((cell.value),version=5,error='H')
    qr.png(r"./qrcodes/"+str(cell.value)+".png",scale = 6)

print("QR Codes successfully generated! along with final.xlsx")


    